"""
Integration tests for src/core/checklist_autofiller.py

Fixture: C160112-020725_Huahong Grace Fab3 #2_NX-Wafer 200mm_Industrial Checklist_2026-Q1-Rev.0.xlsx
         Located at: tests/fixtures/checklist_sample.xlsx  (or set CHECKLIST_FIXTURE env var)
"""

import os
import shutil
from pathlib import Path

import pytest
import openpyxl

# ---------------------------------------------------------------------------
# Fixture resolution
# ---------------------------------------------------------------------------

FIXTURE_ENV = os.environ.get('CHECKLIST_FIXTURE', '')
FIXTURE_CANDIDATES = [
    Path(FIXTURE_ENV) if FIXTURE_ENV else None,
    Path('tests/fixtures/checklist_sample.xlsx'),
    Path(r'C:\Users\Spare\Desktop\C160112-020725_Huahong Grace Fab3 #2_NX-Wafer 200mm_Industrial Checklist_2026-Q1-Rev.0.xlsx'),
]


def _find_fixture() -> Path | None:
    for p in FIXTURE_CANDIDATES:
        if p and p.exists():
            return p
    return None


FIXTURE_PATH = _find_fixture()
needs_fixture = pytest.mark.skipif(
    FIXTURE_PATH is None,
    reason="Checklist Excel fixture not found. "
           "Set CHECKLIST_FIXTURE env var or place file at tests/fixtures/checklist_sample.xlsx"
)

# ---------------------------------------------------------------------------
# Minimal QC report for tests (a handful of known items)
# ---------------------------------------------------------------------------

SAMPLE_QC_REPORT = {
    'results': [
        {'module': 'Dsp', 'part_type': 'ZDetector', 'part_name': 'LongHead',
         'item_name': 'MicrometerPerAdcHigh', 'actual_value': 0.0123},
        {'module': 'Dsp', 'part_type': 'XScanner', 'part_name': '100um',
         'item_name': 'ServoCutoffFrequencyHz', 'actual_value': 800},
        {'module': 'Dsp', 'part_type': 'AFMHead', 'part_name': 'Probe',
         'item_name': 'LaserPowerMw', 'actual_value': 3.5},
    ],
    'summary': {'pass_rate': 100, 'excluded': 0},
}


# ---------------------------------------------------------------------------
# Helper: create a minimal synthetic checklist workbook
# ---------------------------------------------------------------------------

def _make_minimal_checklist(tmp_path: Path) -> Path:
    """Create a minimal checklist Excel with the expected sheet structure."""
    wb = openpyxl.Workbook()
    # 표지 sheet (should be ignored)
    ws_cover = wb.active
    ws_cover.title = '표지'

    # Main sheet
    ws = wb.create_sheet('통합_(NX-Wafer 200mm)')
    headers = ['No', 'Module', 'Check Items', 'Min', 'Criteria', 'Max',
               'Measurement', 'Unit', 'PASS/FAIL', 'Category', '',
               '', 'DB_Key']
    for i, h in enumerate(headers, 1):
        ws.cell(1, i).value = h

    # Row 2: group header (D/E/G = '-')
    ws.cell(2, 2).value = 'XY Stage'
    ws.cell(2, 4).value = '-'
    ws.cell(2, 5).value = '-'
    ws.cell(2, 7).value = '-'

    # Row 3: normal item — explicit DB_Key in M column (should use stage A)
    ws.cell(3, 2).value = 'Z Detector'
    ws.cell(3, 3).value = 'MicrometerPerAdcHigh measurement'
    ws.cell(3, 8).value = 'um/ADC'
    ws.cell(3, 13).value = 'Dsp.ZDetector.LongHead.MicrometerPerAdcHigh'

    # Row 4: item with formula in G (should be protected)
    ws.cell(4, 2).value = 'AFM Head'
    ws.cell(4, 3).value = 'LINEST formula row'
    ws.cell(4, 7).value = '=LINEST(A1:A10)'

    # Row 5: normal item — no explicit key (relies on mapper)
    ws.cell(5, 2).value = 'AFM Head'
    ws.cell(5, 3).value = 'Laser Power'
    ws.cell(5, 8).value = 'mW'

    out = tmp_path / 'checklist_test.xlsx'
    wb.save(str(out))
    return out


# ---------------------------------------------------------------------------
# Tests using synthetic workbook
# ---------------------------------------------------------------------------

class TestAutoFillerSynthetic:

    def test_dry_run_produces_report_no_file_change(self, tmp_path):
        xl = _make_minimal_checklist(tmp_path)
        mtime_before = xl.stat().st_mtime

        from src.core.checklist_autofiller import ChecklistAutoFiller
        filler = ChecklistAutoFiller(
            excel_path=str(xl),
            qc_report=SAMPLE_QC_REPORT,
            dry_run=True,
        )
        report = filler.run()

        # File should NOT be modified in dry-run
        assert xl.stat().st_mtime == mtime_before
        assert report is not None

    def test_dry_run_counts_are_non_negative(self, tmp_path):
        xl = _make_minimal_checklist(tmp_path)

        from src.core.checklist_autofiller import ChecklistAutoFiller
        filler = ChecklistAutoFiller(
            excel_path=str(xl), qc_report=SAMPLE_QC_REPORT, dry_run=True,
        )
        report = filler.run()

        assert report.filled >= 0
        assert report.skipped_unmapped >= 0
        assert report.skipped_protected >= 0

    def test_group_header_row_not_filled(self, tmp_path):
        xl = _make_minimal_checklist(tmp_path)

        from src.core.checklist_autofiller import ChecklistAutoFiller
        filler = ChecklistAutoFiller(
            excel_path=str(xl), qc_report=SAMPLE_QC_REPORT, dry_run=False,
        )
        filler.run()

        wb = openpyxl.load_workbook(str(xl), data_only=False)
        ws = wb['통합_(NX-Wafer 200mm)']
        # Row 2 is group header — G2 should remain '-'
        assert ws.cell(2, 7).value == '-'

    def test_formula_cell_not_overwritten(self, tmp_path):
        xl = _make_minimal_checklist(tmp_path)

        from src.core.checklist_autofiller import ChecklistAutoFiller
        filler = ChecklistAutoFiller(
            excel_path=str(xl), qc_report=SAMPLE_QC_REPORT, dry_run=False,
        )
        filler.run()

        wb = openpyxl.load_workbook(str(xl), data_only=False)
        ws = wb['통합_(NX-Wafer 200mm)']
        # Row 4 has a formula — must be preserved
        assert str(ws.cell(4, 7).value or '').startswith('=')

    def test_cover_sheet_untouched(self, tmp_path):
        xl = _make_minimal_checklist(tmp_path)
        # Note '표지' sheet has no data to overwrite; verify it still exists
        from src.core.checklist_autofiller import ChecklistAutoFiller
        filler = ChecklistAutoFiller(
            excel_path=str(xl), qc_report=SAMPLE_QC_REPORT, dry_run=False,
        )
        filler.run()

        wb = openpyxl.load_workbook(str(xl))
        assert '표지' in wb.sheetnames

    def test_backup_file_created_on_write(self, tmp_path):
        xl = _make_minimal_checklist(tmp_path)

        from src.core.checklist_autofiller import ChecklistAutoFiller
        filler = ChecklistAutoFiller(
            excel_path=str(xl), qc_report=SAMPLE_QC_REPORT, dry_run=False,
        )
        filler.run()

        backups = list(tmp_path.glob('*.bak.*.xlsx'))
        assert len(backups) == 1

    def test_log_sheet_created(self, tmp_path):
        xl = _make_minimal_checklist(tmp_path)

        from src.core.checklist_autofiller import ChecklistAutoFiller
        filler = ChecklistAutoFiller(
            excel_path=str(xl), qc_report=SAMPLE_QC_REPORT, dry_run=False,
        )
        filler.run()

        wb = openpyxl.load_workbook(str(xl))
        assert '_AutoFill_Log' in wb.sheetnames

    def test_explicit_mapping_fills_g_column(self, tmp_path):
        """Row 3 has explicit DB_Key in M → G should receive the QC value."""
        xl = _make_minimal_checklist(tmp_path)

        from src.core.checklist_autofiller import ChecklistAutoFiller
        filler = ChecklistAutoFiller(
            excel_path=str(xl), qc_report=SAMPLE_QC_REPORT,
            dry_run=False, confidence_threshold=0.0,
        )
        filler.run()

        wb = openpyxl.load_workbook(str(xl), data_only=True)
        ws = wb['통합_(NX-Wafer 200mm)']
        # Row 3 should have the QC value 0.0123
        g3 = ws.cell(3, 7).value
        assert g3 == pytest.approx(0.0123, abs=1e-6)

    def test_no_write_to_i_column(self, tmp_path):
        """I-column (PASS/FAIL) must never be written."""
        xl = _make_minimal_checklist(tmp_path)

        from src.core.checklist_autofiller import ChecklistAutoFiller
        filler = ChecklistAutoFiller(
            excel_path=str(xl), qc_report=SAMPLE_QC_REPORT,
            dry_run=False, confidence_threshold=0.0,
        )
        filler.run()

        wb = openpyxl.load_workbook(str(xl), data_only=False)
        ws = wb['통합_(NX-Wafer 200mm)']
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row, 9).value  # column I
            if val is not None:
                # Must not be a plain scalar we wrote — formulas are fine
                assert str(val).startswith('=') or val is None, \
                    f"I-column row {row} unexpectedly has non-formula value: {val!r}"


# ---------------------------------------------------------------------------
# Tests with the real fixture file
# ---------------------------------------------------------------------------

class TestAutoFillerWithFixture:

    @needs_fixture
    def test_dry_run_processes_all_rows_without_error(self, tmp_path):
        """Given the real checklist fixture, dry-run must complete without exception."""
        xl = tmp_path / 'checklist_fixture.xlsx'
        shutil.copy2(str(FIXTURE_PATH), str(xl))

        from src.core.checklist_autofiller import ChecklistAutoFiller
        filler = ChecklistAutoFiller(
            excel_path=str(xl), qc_report=SAMPLE_QC_REPORT, dry_run=True,
        )
        report = filler.run()
        assert report.total_processed > 0

    @needs_fixture
    def test_row_counts_sum_to_total(self, tmp_path):
        """filled + skipped_unmapped + skipped_protected + group_headers == total_processed."""
        xl = tmp_path / 'checklist_fixture.xlsx'
        shutil.copy2(str(FIXTURE_PATH), str(xl))

        from src.core.checklist_autofiller import ChecklistAutoFiller
        filler = ChecklistAutoFiller(
            excel_path=str(xl), qc_report=SAMPLE_QC_REPORT, dry_run=True,
        )
        report = filler.run()
        assert report.total_processed == (
            report.filled + report.skipped_unmapped +
            report.skipped_protected + report.skipped_group_header
        )

    @needs_fixture
    def test_explicit_row_m119_mapped(self, tmp_path):
        """Row 119 has the only pre-filled DB_Key in the fixture — should be filled."""
        xl = tmp_path / 'checklist_fixture.xlsx'
        shutil.copy2(str(FIXTURE_PATH), str(xl))

        qc = {
            'results': [{
                'module': 'Dsp', 'part_type': 'ZDetector',
                'part_name': 'LongHead', 'item_name': 'MicrometerPerAdcHigh',
                'actual_value': 0.0123,
            }],
            'summary': {'pass_rate': 100, 'excluded': 0},
        }

        from src.core.checklist_autofiller import ChecklistAutoFiller
        filler = ChecklistAutoFiller(
            excel_path=str(xl), qc_report=qc, dry_run=True,
        )
        report = filler.run()

        explicit_rows = [e for e in report.mapping_changes if e['source'] == 'explicit']
        assert any(e['row'] == 119 for e in explicit_rows), \
            "Row 119 (the only pre-filled DB_Key row) was not mapped as 'explicit'"
