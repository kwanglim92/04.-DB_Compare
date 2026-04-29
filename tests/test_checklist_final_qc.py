from pathlib import Path

import openpyxl
import pytest

from src.core.checklist_final_qc import (
    ACTION_FILL,
    ACTION_DO_NOT_WRITE,
    ACTION_REPLACE,
    ACTION_REVIEW,
    PREFLIGHT_ERROR,
    PREFLIGHT_INFO,
    PREFLIGHT_WARNING,
    RISK_BLOCKED,
    RISK_HIGH,
    RISK_SAFE,
    STATUS_MISMATCH,
    STATUS_MISSING,
    STATUS_NON_COMPARABLE,
    STATUS_OK,
    STATUS_PROTECTED,
    STATUS_SKIPPED_GROUP,
    STATUS_UNMAPPED,
    ChecklistFinalQcEngine,
)


SAMPLE_QC_REPORT = {
    "results": [
        {
            "module": "Dsp",
            "part_type": "ZDetector",
            "part_name": "LongHead",
            "item_name": "MicrometerPerAdcHigh",
            "actual_value": 0.0123,
        },
        {
            "module": "Dsp",
            "part_type": "XScanner",
            "part_name": "100um",
            "item_name": "ServoCutoffFrequencyHz",
            "actual_value": 800,
        },
        {
            "module": "Dsp",
            "part_type": "AFMHead",
            "part_name": "Probe",
            "item_name": "LaserPowerMw",
            "actual_value": 3.5,
        },
    ],
    "summary": {"pass_rate": 100},
}


def make_checklist(tmp_path: Path) -> Path:
    wb = openpyxl.Workbook()
    cover = wb.active
    cover.title = "표지"
    ws = wb.create_sheet("통합_(NX-Wafer 200mm)")

    headers = [
        "No",
        "Module",
        "Check Items",
        "Min",
        "Criteria",
        "Max",
        "Measurement",
        "Unit",
        "PASS/FAIL",
        "Category",
        "",
        "",
        "DB_Key",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(1, col).value = header

    # Group header: skipped.
    ws.cell(2, 2).value = "XY Stage"
    ws.cell(2, 4).value = "-"
    ws.cell(2, 5).value = "-"
    ws.cell(2, 7).value = "-"

    # OK: current G equals QC value.
    ws.cell(3, 2).value = "Z Detector"
    ws.cell(3, 3).value = "MicrometerPerAdcHigh"
    ws.cell(3, 7).value = 0.0123
    ws.cell(3, 13).value = "Dsp.ZDetector.LongHead.MicrometerPerAdcHigh"

    # Missing: blank G with explicit mapping.
    ws.cell(4, 2).value = "XY Stage"
    ws.cell(4, 3).value = "Servo cutoff"
    ws.cell(4, 13).value = "Dsp.XScanner.100um.ServoCutoffFrequencyHz"

    # Mismatch: existing G differs from QC.
    ws.cell(5, 2).value = "XY Stage"
    ws.cell(5, 3).value = "Servo cutoff wrong"
    ws.cell(5, 7).value = 700
    ws.cell(5, 13).value = "Dsp.XScanner.100um.ServoCutoffFrequencyHz"

    # Protected: formula in G.
    ws.cell(6, 2).value = "AFM Head"
    ws.cell(6, 3).value = "Laser Power"
    ws.cell(6, 7).value = "=SUM(1,2)"
    ws.cell(6, 13).value = "Dsp.AFMHead.Probe.LaserPowerMw"

    # Unmapped/manual-learning candidate.
    ws.cell(7, 2).value = "AFM Head"
    ws.cell(7, 3).value = "Human Approved Customer Label"

    path = tmp_path / "checklist.xlsx"
    wb.save(path)
    return path


def rows_by_number(report):
    return {row.row: row for row in report.rows}


def issue_codes(preflight_result):
    return {issue.code: issue.severity for issue in preflight_result.issues}


def test_analyze_classifies_final_checklist_rows(tmp_path):
    checklist = make_checklist(tmp_path)
    engine = ChecklistFinalQcEngine(mapping_dir=tmp_path / "mappings")

    report = engine.analyze(str(checklist), SAMPLE_QC_REPORT, "Fallback Profile")
    rows = rows_by_number(report)

    assert report.model == "NX-Wafer 200mm"
    assert rows[2].status == STATUS_SKIPPED_GROUP
    assert rows[3].status == STATUS_OK
    assert rows[4].status == STATUS_MISSING
    assert rows[5].status == STATUS_MISMATCH
    assert rows[6].status == STATUS_PROTECTED
    assert rows[7].status == STATUS_UNMAPPED
    assert rows[4].risk_level == RISK_SAFE
    assert rows[4].recommended_action == ACTION_FILL
    assert rows[5].risk_level == RISK_SAFE
    assert rows[5].recommended_action == ACTION_REPLACE
    assert rows[5].delta == "+100"
    assert rows[6].risk_level == RISK_BLOCKED
    assert rows[6].recommended_action == ACTION_DO_NOT_WRITE
    assert rows[7].risk_level == RISK_BLOCKED

    safe_rows = [row.row for row in report.rows if row.is_safe_correction]
    assert safe_rows == [4, 5]


def test_preflight_reports_workbook_readiness(tmp_path):
    checklist = make_checklist(tmp_path)
    engine = ChecklistFinalQcEngine(mapping_dir=tmp_path / "mappings")

    result = engine.preflight(str(checklist), SAMPLE_QC_REPORT)

    assert not result.has_errors
    assert result.main_sheet == "통합_(NX-Wafer 200mm)"
    assert result.db_key_column == 13
    assert result.qc_key_count == 3
    assert issue_codes(result)["FORMULA_CELLS"] == PREFLIGHT_INFO


def test_preflight_blocks_missing_file_and_empty_qc_lookup(tmp_path):
    engine = ChecklistFinalQcEngine(mapping_dir=tmp_path / "mappings")

    missing = engine.preflight(str(tmp_path / "missing.xlsx"), SAMPLE_QC_REPORT)
    missing_codes = issue_codes(missing)
    assert missing.has_errors
    assert missing_codes["FILE_NOT_FOUND"] == PREFLIGHT_ERROR

    checklist = make_checklist(tmp_path)
    empty_qc = engine.preflight(str(checklist), {"results": []})
    empty_codes = issue_codes(empty_qc)
    assert empty_qc.has_errors
    assert empty_codes["QC_LOOKUP_EMPTY"] == PREFLIGHT_ERROR


def test_preflight_flags_sheet_name_structure_and_features(tmp_path):
    checklist = make_checklist(tmp_path)
    wb = openpyxl.load_workbook(checklist)
    ws = wb["통합_(NX-Wafer 200mm)"]
    ws.title = "Checklist"
    ws.cell(1, 13).value = "Internal Notes"
    for row in range(2, 8):
        ws.cell(row, 13).value = None
    ws.row_dimensions[5].hidden = True
    ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=3)
    ws.protection.sheet = True
    wb.save(checklist)

    engine = ChecklistFinalQcEngine(mapping_dir=tmp_path / "mappings")
    result = engine.preflight(str(checklist), SAMPLE_QC_REPORT)
    codes = issue_codes(result)

    assert not result.has_errors
    assert result.main_sheet == "Checklist"
    assert codes["MAIN_SHEET_NAME"] == PREFLIGHT_WARNING
    assert codes["DB_KEY_COLUMN"] == PREFLIGHT_WARNING
    assert codes["HIDDEN_ROWS"] == PREFLIGHT_WARNING
    assert codes["MERGED_CELLS"] == PREFLIGHT_WARNING
    assert codes["SHEET_PROTECTION"] == PREFLIGHT_WARNING


def test_apply_approved_writes_only_to_copy_and_rechecks(tmp_path):
    checklist = make_checklist(tmp_path)
    original_mtime = checklist.stat().st_mtime
    output = tmp_path / "filled_copy.xlsx"
    engine = ChecklistFinalQcEngine(mapping_dir=tmp_path / "mappings")
    report = engine.analyze(str(checklist), SAMPLE_QC_REPORT, "")

    result = engine.apply_approved(report, [4, 5, 7], str(output), write_db_key=False)

    assert result.applied_count == 2
    assert result.skipped_count == 1
    assert checklist.stat().st_mtime == pytest.approx(original_mtime)

    original_wb = openpyxl.load_workbook(checklist, data_only=False)
    original_ws = original_wb["통합_(NX-Wafer 200mm)"]
    assert original_ws.cell(4, 7).value is None
    assert original_ws.cell(5, 7).value == 700
    assert "_AutoFill_Log" not in original_wb.sheetnames

    out_wb = openpyxl.load_workbook(output, data_only=False)
    out_ws = out_wb["통합_(NX-Wafer 200mm)"]
    assert out_ws.cell(2, 7).value == "-"
    assert str(out_ws.cell(6, 7).value).startswith("=")
    assert out_ws.cell(4, 7).value == 800
    assert out_ws.cell(5, 7).value == 800
    assert out_ws.cell(4, 13).value == "Dsp.XScanner.100um.ServoCutoffFrequencyHz"
    assert "_AutoFill_Log" in out_wb.sheetnames
    log_ws = out_wb["_AutoFill_Log"]
    headers = [log_ws.cell(1, col).value for col in range(1, log_ws.max_column + 1)]
    assert {
        "Row",
        "StatusBefore",
        "StatusAfter",
        "Risk",
        "RecommendedAction",
        "ReviewerDecision",
        "ExceptionReason",
        "Module",
        "Item",
        "DB_Key",
        "Source",
        "Confidence",
        "PreviousG",
        "QCValue",
        "WriteDBKey",
        "Timestamp",
    }.issubset(set(headers))

    post_rows = rows_by_number(result.post_report)
    assert post_rows[4].status == STATUS_OK
    assert post_rows[5].status == STATUS_OK


def test_default_output_path_uses_internal_copy_suffix(tmp_path):
    checklist = make_checklist(tmp_path)
    engine = ChecklistFinalQcEngine(mapping_dir=tmp_path / "mappings")

    output = Path(engine.default_output_path(str(checklist)))

    assert output.parent == checklist.parent
    assert output.name.startswith("checklist_QC_Checked_AutoFilled_")
    assert output.suffix == ".xlsx"


def test_apply_default_writes_db_key_for_exact_safe_row(tmp_path):
    checklist = make_checklist(tmp_path)
    wb = openpyxl.load_workbook(checklist)
    ws = wb["통합_(NX-Wafer 200mm)"]
    ws.cell(8, 2).value = "AFM Head"
    ws.cell(8, 3).value = "LaserPowerMw"
    wb.save(checklist)

    output = tmp_path / "exact_filled.xlsx"
    engine = ChecklistFinalQcEngine(mapping_dir=tmp_path / "mappings")
    report = engine.analyze(str(checklist), SAMPLE_QC_REPORT, "")
    row = report.get_row(8)
    assert row.status == STATUS_MISSING
    assert row.risk_level == RISK_SAFE

    result = engine.apply_approved(report, [8], str(output))

    assert result.applied_count == 1
    out_wb = openpyxl.load_workbook(output, data_only=False)
    out_ws = out_wb["통합_(NX-Wafer 200mm)"]
    assert out_ws.cell(8, 7).value == 3.5
    assert out_ws.cell(8, 13).value == "Dsp.AFMHead.Probe.LaserPowerMw"


def test_apply_skips_blocked_rows_even_when_approved(tmp_path):
    checklist = make_checklist(tmp_path)
    output = tmp_path / "blocked_rows.xlsx"
    engine = ChecklistFinalQcEngine(mapping_dir=tmp_path / "mappings")
    report = engine.analyze(str(checklist), SAMPLE_QC_REPORT, "")

    result = engine.apply_approved(report, [2, 6, 7], str(output))

    assert result.applied_count == 0
    assert result.skipped_count == 3
    out_wb = openpyxl.load_workbook(output, data_only=False)
    out_ws = out_wb["통합_(NX-Wafer 200mm)"]
    assert out_ws.cell(2, 7).value == "-"
    assert str(out_ws.cell(6, 7).value).startswith("=")
    assert out_ws.cell(7, 7).value is None


def test_high_risk_mismatch_requires_reviewer_confirmation(tmp_path):
    checklist = make_checklist(tmp_path)
    wb = openpyxl.load_workbook(checklist)
    ws = wb["통합_(NX-Wafer 200mm)"]
    ws.cell(5, 7).value = "wrong text"
    wb.save(checklist)

    engine = ChecklistFinalQcEngine(mapping_dir=tmp_path / "mappings")
    report = engine.analyze(str(checklist), SAMPLE_QC_REPORT, "")
    row = report.get_row(5)

    assert row.status == STATUS_MISMATCH
    assert row.risk_level == RISK_HIGH
    assert row.recommended_action == ACTION_REVIEW
    assert row.requires_review


def test_qc_na_is_non_comparable_and_blocked(tmp_path):
    checklist = make_checklist(tmp_path)
    qc_report = {
        "results": [
            {
                "module": "Dsp",
                "part_type": "XScanner",
                "part_name": "100um",
                "item_name": "ServoCutoffFrequencyHz",
                "actual_value": "N/A",
            },
        ]
    }
    engine = ChecklistFinalQcEngine(mapping_dir=tmp_path / "mappings")

    report = engine.analyze(str(checklist), qc_report, "")
    row = report.get_row(4)

    assert row.status == STATUS_NON_COMPARABLE
    assert row.risk_level == RISK_BLOCKED
    assert row.recommended_action == ACTION_DO_NOT_WRITE


def test_mismatch_exception_is_logged_without_changing_measurement(tmp_path):
    checklist = make_checklist(tmp_path)
    output = tmp_path / "exception_logged.xlsx"
    engine = ChecklistFinalQcEngine(mapping_dir=tmp_path / "mappings")
    report = engine.analyze(str(checklist), SAMPLE_QC_REPORT, "")

    result = engine.apply_approved(report, [], str(output), exceptions={5: "Customer accepted previous value"})

    assert result.applied_count == 0
    assert result.exception_count == 1
    out_wb = openpyxl.load_workbook(output, data_only=False)
    out_ws = out_wb["통합_(NX-Wafer 200mm)"]
    assert out_ws.cell(5, 7).value == 700

    log_ws = out_wb["_AutoFill_Log"]
    headers = [log_ws.cell(1, col).value for col in range(1, log_ws.max_column + 1)]
    decision_col = headers.index("ReviewerDecision") + 1
    reason_col = headers.index("ExceptionReason") + 1
    assert log_ws.cell(2, decision_col).value == "Keep Checklist Value / Exception"
    assert log_ws.cell(2, reason_col).value == "Customer accepted previous value"


def test_manual_approved_mapping_is_learned_for_next_analysis(tmp_path):
    checklist = make_checklist(tmp_path)
    output = tmp_path / "manual_filled.xlsx"
    engine = ChecklistFinalQcEngine(mapping_dir=tmp_path / "mappings")
    report = engine.analyze(str(checklist), SAMPLE_QC_REPORT, "")
    manual_row = report.get_row(7)

    manual_row.db_key = "Dsp.AFMHead.Probe.LaserPowerMw"
    manual_row.qc_value = 3.5
    manual_row.source = "manual"
    manual_row.confidence = 0.95
    manual_row.status = engine.classify_value(manual_row.current_value, manual_row.qc_value)

    result = engine.apply_approved(report, [7], str(output), write_db_key=False)
    assert result.applied_count == 1
    assert result.learned_count == 1

    next_report = engine.analyze(str(checklist), SAMPLE_QC_REPORT, "")
    learned_row = next_report.get_row(7)
    assert learned_row.source == "learned"
    assert learned_row.db_key == "Dsp.AFMHead.Probe.LaserPowerMw"
    assert learned_row.status == STATUS_MISSING
