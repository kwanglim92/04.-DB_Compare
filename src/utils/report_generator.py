"""
Excel Report Generator
Generates professional QC inspection reports in Excel format
"""

from datetime import datetime
from pathlib import Path
from typing import Dict, Optional
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter

from src.constants import EXCEL_COLORS
from src.utils.format_helpers import format_spec
from src.utils.template_helper import (
    get_default_template, load_template, validate_template, apply_placeholders
)

logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    """Generate Excel reports from QC inspection results."""

    def __init__(self, template: Optional[Dict] = None):
        self.logger = logger
        self.colors = EXCEL_COLORS

        if template is None:
            self._template = get_default_template()
        else:
            self._template = validate_template(template)

    @classmethod
    def from_template_file(cls, template_path: Path) -> "ExcelReportGenerator":
        """Convenience constructor that loads template from a JSON file."""
        tmpl = load_template(template_path)
        return cls(template=tmpl)

    # ------------------------------------------------------------------
    # Header color helper
    # ------------------------------------------------------------------

    def _header_fill(self) -> PatternFill:
        color = self._template.get("header_color", "#1f6aa5").lstrip("#")
        return PatternFill(start_color=color, fill_type="solid")

    def _footer_text(self) -> str:
        return self._template.get("footer_text", "")

    def _report_title(self, qc_report: Dict) -> str:
        title_tpl = self._template.get("title_template", "QC Inspection Report — {profile} — {date}")
        ctx = {
            "profile": qc_report.get("profile_name", ""),
            "date": datetime.now().strftime("%Y-%m-%d"),
            "engineer": self._template.get("engineer_name", ""),
            "instrument": qc_report.get("instrument", ""),
        }
        return apply_placeholders(title_tpl, ctx)

    # ------------------------------------------------------------------
    # Public entry point
    # ------------------------------------------------------------------

    def generate_report(self, qc_report: Dict, output_path: str) -> bool:
        """Generate complete Excel report.

        Args:
            qc_report: QC report from Comparator
            output_path: Path to save Excel file
        """
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        sheets = self._template.get("sheets", {})

        if sheets.get("cover_page", False):
            self.create_cover_sheet(wb, qc_report)
        if sheets.get("summary", True):
            self.create_summary_sheet(wb, qc_report)
        if sheets.get("all_items", True):
            self.create_all_items_sheet(wb, qc_report)
        if sheets.get("failed_items", True):
            self.create_failed_items_sheet(wb, qc_report)

        # Safety: if nothing was created, add Summary
        if not wb.sheetnames:
            self.create_summary_sheet(wb, qc_report)

        wb.save(output_path)
        self.logger.info(f"Report saved to: {output_path}")
        return True

    
    def create_cover_sheet(self, wb: Workbook, report: Dict):
        """Create a cover page sheet with company info and inspection details."""
        ws = wb.create_sheet("Cover", 0)
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 40

        company = self._template.get("company", {})
        logo_path = company.get("logo_path", "")

        # Logo
        if logo_path:
            try:
                from openpyxl.drawing.image import Image as XlImage
                img = XlImage(logo_path)
                img.width = min(img.width, 200)
                img.height = min(img.height, 80)
                ws.add_image(img, "A1")
                ws.row_dimensions[1].height = 60
            except Exception as e:
                self.logger.warning(f"Cannot insert logo: {e}")

        row = 4
        ws[f'A{row}'] = self._report_title(report)
        ws[f'A{row}'].font = Font(size=16, bold=True)
        ws.merge_cells(f'A{row}:B{row}')
        row += 2

        info = [
            ("Company", company.get("name", "")),
            ("Engineer", self._template.get("engineer_name", "")),
            ("Instrument", report.get("instrument", "")),
            ("Profile", report.get("profile_name", "")),
            ("Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("Contact", company.get("contact", "")),
        ]
        for label, value in info:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1

        footer = self._footer_text()
        if footer:
            row += 2
            ws[f'A{row}'] = footer
            ws[f'A{row}'].font = Font(italic=True, color="808080")
            ws.merge_cells(f'A{row}:B{row}')

    def create_summary_sheet(self, wb: Workbook, report: Dict):
        """Create summary sheet with statistics"""
        ws = wb.create_sheet("Summary")

        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30

        # Title from template
        ws['A1'] = self._report_title(report)
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:B1')
        
        # Metadata
        row = 3
        metadata = [
            ("Profile:", report.get('profile_name', '')),
            ("Timestamp:", report.get('timestamp', '')),
            ("DB Root:", report.get('db_root', '')),
            ("Instrument:", report.get('instrument', ''))
        ]
        
        for label, value in metadata:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        # Statistics
        row += 2
        ws[f'A{row}'] = "Statistics"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        summary = report.get('summary', {})
        stats = [
            ("Total Items:", summary.get('total_items', 0)),
            ("Validated Items:", summary.get('validated', 0)),
            ("Checked Only:", summary.get('checked_only', 0)),
            ("✓ Passed:", summary.get('passed', 0)),
            ("✗ Failed:", summary.get('failed', 0)),
            ("○ No Spec:", summary.get('no_spec', 0)),
            ("⚠ Errors:", summary.get('errors', 0)),
            ("", ""),
            ("Pass Rate:", f"{summary.get('pass_rate', 0)}%")
        ]
        
        for label, value in stats:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            
            if "Passed" in label:
                ws[f'A{row}'].fill = PatternFill(start_color=self.colors['pass_bg'], fill_type="solid")
            elif "Failed" in label:
                ws[f'A{row}'].fill = PatternFill(start_color=self.colors['fail_bg'], fill_type="solid")
            elif "Pass Rate" in label:
                ws[f'A{row}'].font = Font(bold=True, size=12)
                ws[f'B{row}'].font = Font(bold=True, size=12, color="006400")
            
            ws[f'B{row}'] = value
            row += 1
    
    def create_all_items_sheet(self, wb: Workbook, report: Dict):
        """Create sheet with all inspection items"""
        ws = wb.create_sheet("All Items")
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 40
        
        # Headers
        headers = ['Module', 'Part Type', 'Part Name', 'Item Name', 
                   'Actual Value', 'Spec', 'Status', 'Message']
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = self._header_fill()
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data
        row = 2
        for result in report.get('results', []):
            ws[f'A{row}'] = result.get('module', '')
            ws[f'B{row}'] = result.get('part_type', '')
            ws[f'C{row}'] = result.get('part_name', '')
            ws[f'D{row}'] = result.get('item_name', '')
            ws[f'E{row}'] = result.get('actual_value', '')
            ws[f'F{row}'] = self.format_spec(result.get('spec'))
            ws[f'G{row}'] = result.get('status', '')
            ws[f'H{row}'] = result.get('message', '')
            
            # Color code status
            status = result.get('status', '')
            if status == 'PASS':
                ws[f'G{row}'].fill = PatternFill(start_color=self.colors['pass_bg'], fill_type="solid")
            elif status == 'FAIL':
                ws[f'G{row}'].fill = PatternFill(start_color=self.colors['fail_bg'], fill_type="solid")
                ws[f'G{row}'].font = Font(bold=True, color="8B0000")
            elif status == 'CHECK':
                ws[f'G{row}'].fill = PatternFill(start_color=self.colors['check_bg'], fill_type="solid")
                ws[f'G{row}'].font = Font(color="1976D2")
            elif status == 'NO_SPEC':
                ws[f'G{row}'].fill = PatternFill(start_color="FFFFF9C4", fill_type="solid")  # light yellow
            elif status == 'ERROR':
                ws[f'G{row}'].fill = PatternFill(start_color="FFFCE4EC", fill_type="solid")  # light pink
            
            ws[f'G{row}'].alignment = Alignment(horizontal='center')
            row += 1
        
        # Freeze first row
        ws.freeze_panes = 'A2'
    
    def create_failed_items_sheet(self, wb: Workbook, report: Dict):
        """Create sheet with failed items only"""
        ws = wb.create_sheet("Failed Items")
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 40
        
        # Headers
        headers = ['Module', 'Part Type', 'Part Name', 'Item Name', 
                   'Actual Value', 'Spec', 'Status', 'Message']
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = self._header_fill()
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Filter failed items
        failed_items = [r for r in report.get('results', []) if r.get('status') == 'FAIL']
        
        if not failed_items:
            ws['A2'] = "No failed items"
            ws['A2'].font = Font(italic=True, color="808080")
            return
        
        # Data
        row = 2
        for result in failed_items:
            ws[f'A{row}'] = result.get('module', '')
            ws[f'B{row}'] = result.get('part_type', '')
            ws[f'C{row}'] = result.get('part_name', '')
            ws[f'D{row}'] = result.get('item_name', '')
            ws[f'E{row}'] = result.get('actual_value', '')
            ws[f'F{row}'] = self.format_spec(result.get('spec'))
            ws[f'G{row}'] = result.get('status', '')
            ws[f'H{row}'] = result.get('message', '')
            
            # Highlight failed items
            ws[f'G{row}'].fill = PatternFill(start_color=self.colors['fail_bg'], fill_type="solid")
            ws[f'G{row}'].font = Font(bold=True, color="8B0000")
            ws[f'G{row}'].alignment = Alignment(horizontal='center')
            
            # Highlight actual value that failed
            ws[f'E{row}'].font = Font(bold=True, color="8B0000")
            
            row += 1
        
        # Freeze first row
        ws.freeze_panes = 'A2'
    
    def format_spec(self, spec: Dict) -> str:
        """Forward to common format helper"""
        return format_spec(spec)
