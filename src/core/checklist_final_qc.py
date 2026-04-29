"""
Final Checklist QC engine.

The engine first analyzes a completed customer checklist without modifying it,
then writes only reviewer-approved rows to a timestamped copy.
"""

import fnmatch
import json
import logging
import shutil
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional

import openpyxl
from openpyxl.styles import PatternFill

from src.constants import CHECKLIST_COLS
from src.core.checklist_mapper import ChecklistMapper
from src.core.checklist_validator import ChecklistValidator
from src.utils.config_helper import get_cache_dir
from src.utils.text_normalizer import normalize

logger = logging.getLogger(__name__)

STATUS_OK = "OK"
STATUS_MISSING = "Missing"
STATUS_MISMATCH = "Mismatch"
STATUS_UNMAPPED = "Unmapped"
STATUS_PROTECTED = "Protected"
STATUS_SKIPPED_GROUP = "SkippedGroup"
STATUS_NON_COMPARABLE = "NonComparable"

RISK_SAFE = "Safe"
RISK_REVIEW = "Review"
RISK_HIGH = "HighRisk"
RISK_BLOCKED = "Blocked"
RISK_OK = "OK"

ACTION_FILL = "Fill QC Value"
ACTION_REPLACE = "Replace with QC Value"
ACTION_REVIEW = "Reviewer Confirm"
ACTION_DO_NOT_WRITE = "Do Not Write"
ACTION_NONE = "No Action"

PREFLIGHT_ERROR = "Error"
PREFLIGHT_WARNING = "Warning"
PREFLIGHT_INFO = "Info"

# Column indices — sourced from src.constants.CHECKLIST_COLS
COL_MODULE = CHECKLIST_COLS['MODULE']
COL_ITEM = CHECKLIST_COLS['ITEM']
COL_MIN = CHECKLIST_COLS['MIN']
COL_CRITERIA = CHECKLIST_COLS['CRITERIA']
COL_VALUE = CHECKLIST_COLS['VALUE']
COL_UNIT = CHECKLIST_COLS['UNIT']
COL_DB_KEY = CHECKLIST_COLS['DB_KEY']

_FILL_APPROVED = PatternFill(fill_type="solid", fgColor="FFE8F5E9")
_FILL_MISMATCH = PatternFill(fill_type="solid", fgColor="FFFFF8E1")


@dataclass
class PreflightIssue:
    severity: str
    code: str
    message: str
    sheet_name: str = ""
    row: Optional[int] = None
    column: Optional[int] = None


@dataclass
class PreflightResult:
    excel_path: str
    main_sheet: str = ""
    db_key_column: int = COL_DB_KEY
    qc_key_count: int = 0
    issues: List[PreflightIssue] = field(default_factory=list)

    @property
    def errors(self) -> List[PreflightIssue]:
        return [issue for issue in self.issues if issue.severity == PREFLIGHT_ERROR]

    @property
    def warnings(self) -> List[PreflightIssue]:
        return [issue for issue in self.issues if issue.severity == PREFLIGHT_WARNING]

    @property
    def infos(self) -> List[PreflightIssue]:
        return [issue for issue in self.issues if issue.severity == PREFLIGHT_INFO]

    @property
    def has_errors(self) -> bool:
        return bool(self.errors)

    @property
    def summary(self) -> Dict[str, int]:
        return {
            PREFLIGHT_ERROR: len(self.errors),
            PREFLIGHT_WARNING: len(self.warnings),
            PREFLIGHT_INFO: len(self.infos),
        }


@dataclass
class FinalQcRow:
    row: int
    module: str
    item: str
    current_value: object
    qc_value: object
    db_key: Optional[str]
    source: str
    confidence: float
    status: str
    candidates: List[str] = field(default_factory=list)
    protected_reason: str = ""
    risk_level: str = ""
    recommended_action: str = ""
    auto_approval_reason: str = ""
    requires_review: bool = False
    delta: str = ""

    @property
    def is_writable(self) -> bool:
        return (
            self.status in (STATUS_MISSING, STATUS_MISMATCH)
            and bool(self.db_key)
            and self.risk_level != RISK_BLOCKED
        )

    @property
    def is_safe_correction(self) -> bool:
        return self.is_writable and self.risk_level == RISK_SAFE


@dataclass
class FinalQcReport:
    excel_path: str
    sheet_name: str
    model: str
    qc_report: Dict
    rows: List[FinalQcRow]
    generated_at: str
    preflight_result: Optional[PreflightResult] = None

    @property
    def summary(self) -> Dict[str, int]:
        counts = {
            STATUS_OK: 0,
            STATUS_MISSING: 0,
            STATUS_MISMATCH: 0,
            STATUS_UNMAPPED: 0,
            STATUS_PROTECTED: 0,
            STATUS_SKIPPED_GROUP: 0,
            STATUS_NON_COMPARABLE: 0,
        }
        for row in self.rows:
            counts[row.status] = counts.get(row.status, 0) + 1
        counts["Total"] = len(self.rows)
        return counts

    def get_row(self, row_number: int) -> Optional[FinalQcRow]:
        for row in self.rows:
            if row.row == row_number:
                return row
        return None


@dataclass
class ApplyResult:
    output_path: str
    applied_count: int
    skipped_count: int
    learned_count: int
    post_report: FinalQcReport
    exception_count: int = 0


class ChecklistFinalQcEngine:
    """Read-only final QC plus approved-row writeback to a copied workbook."""

    def __init__(
        self,
        sync_manager=None,
        mapping_dir: Optional[Path] = None,
        confidence_threshold: float = 0.80,
        tolerance: float = 0.001,
    ):
        self.sync_manager = sync_manager
        self.mapping_dir = Path(mapping_dir) if mapping_dir else get_cache_dir()
        self.confidence_threshold = confidence_threshold
        self.tolerance = tolerance
        self._validator = ChecklistValidator()

    def analyze(self, excel_path: str, qc_report: Dict, model: str = "") -> FinalQcReport:
        wb = openpyxl.load_workbook(str(excel_path), data_only=False)
        ws = self._find_main_sheet(wb)
        if ws is None:
            raise ValueError("Checklist main sheet not found.")

        resolved_model = self._extract_model_from_sheet(ws.title) or model or ""
        qc_lookup = self._validator._build_qc_lookup(qc_report)
        db_key_col = self._find_db_key_column(ws)
        raw_rows = self._collect_rows(ws, db_key_col)

        mapper = ChecklistMapper(
            qc_lookup=qc_lookup,
            learned_mappings=self._load_learned_mappings(resolved_model),
            model=resolved_model,
            fuzzy_threshold=self.confidence_threshold,
        )
        map_results = mapper.map_rows(raw_rows)

        rows = []
        for raw, mapped in zip(raw_rows, map_results):
            mapped.row = raw["row"]
            rows.append(self._build_final_row(ws, raw, mapped, qc_lookup))

        return FinalQcReport(
            excel_path=str(excel_path),
            sheet_name=ws.title,
            model=resolved_model,
            qc_report=qc_report,
            rows=rows,
            generated_at=datetime.now().isoformat(timespec="seconds"),
        )

    def preflight(self, excel_path: str, qc_report: Optional[Dict] = None) -> PreflightResult:
        """Inspect workbook readiness before analysis or writeback."""
        result = PreflightResult(excel_path=str(excel_path))
        path = Path(excel_path)

        if not path.exists():
            result.issues.append(PreflightIssue(PREFLIGHT_ERROR, "FILE_NOT_FOUND", "Checklist file was not found."))
            return result
        if path.suffix.lower() not in (".xlsx", ".xlsm"):
            result.issues.append(
                PreflightIssue(PREFLIGHT_WARNING, "FILE_TYPE", "Checklist should be an .xlsx or .xlsm workbook.")
            )

        try:
            with open(path, "rb"):
                pass
        except Exception as exc:
            result.issues.append(
                PreflightIssue(PREFLIGHT_ERROR, "SOURCE_READ", f"Checklist file cannot be read: {exc}")
            )
            return result

        self._check_output_directory(path, result)

        try:
            wb = openpyxl.load_workbook(str(path), data_only=False)
        except Exception as exc:
            result.issues.append(
                PreflightIssue(PREFLIGHT_ERROR, "WORKBOOK_OPEN", f"Workbook cannot be opened: {exc}")
            )
            return result

        integrated_sheets = [name for name in wb.sheetnames if fnmatch.fnmatch(name, "통합_*")]
        if not integrated_sheets:
            result.issues.append(
                PreflightIssue(
                    PREFLIGHT_WARNING,
                    "MAIN_SHEET_NAME",
                    "No 통합_* sheet was found; the first non-cover sheet will be used.",
                )
            )

        ws = self._find_main_sheet(wb)
        if ws is None:
            result.issues.append(
                PreflightIssue(PREFLIGHT_ERROR, "MAIN_SHEET_MISSING", "Checklist main sheet was not found.")
            )
            return result

        result.main_sheet = ws.title
        detected_db_key_col = self._validator._find_db_key_column(ws)
        result.db_key_column = detected_db_key_col or COL_DB_KEY
        if detected_db_key_col is None:
            result.issues.append(
                PreflightIssue(
                    PREFLIGHT_WARNING,
                    "DB_KEY_COLUMN",
                    "DB_Key header was not detected; M column will be used.",
                    sheet_name=ws.title,
                    column=COL_DB_KEY,
                )
            )

        self._check_required_columns(ws, result)
        self._check_sheet_features(ws, result, result.db_key_column)

        if qc_report is not None:
            qc_lookup = self._validator._build_qc_lookup(qc_report)
            result.qc_key_count = len(qc_lookup)
            if not qc_lookup:
                result.issues.append(
                    PreflightIssue(PREFLIGHT_ERROR, "QC_LOOKUP_EMPTY", "QC report has no comparable DB values.")
                )

        return result

    def apply_approved(
        self,
        report: FinalQcReport,
        approved_rows: Iterable[int],
        output_path: str,
        write_db_key: bool = True,
        exceptions: Optional[Dict[int, str]] = None,
    ) -> ApplyResult:
        approved_set = {int(row) for row in approved_rows}
        exception_map = {int(row): str(reason) for row, reason in (exceptions or {}).items()}
        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(report.excel_path, output)

        wb = openpyxl.load_workbook(str(output), data_only=False)
        ws = wb[report.sheet_name] if report.sheet_name in wb.sheetnames else self._find_main_sheet(wb)
        if ws is None:
            raise ValueError("Checklist main sheet not found in output copy.")

        db_key_col = self._find_db_key_column(ws)
        log_entries = []
        applied = 0
        skipped = 0
        learned_rows = []
        timestamp = datetime.now().isoformat(timespec="seconds")

        for row_number, reason in sorted(exception_map.items()):
            final_row = report.get_row(row_number)
            if not final_row:
                skipped += 1
                continue
            self.apply_risk_metadata(final_row)
            previous = ws.cell(row_number, COL_VALUE).value
            log_entries.append(self._log_entry(
                row_number=row_number,
                final_row=final_row,
                previous=previous,
                status_after=final_row.status,
                reviewer_decision="Keep Checklist Value / Exception",
                exception_reason=reason,
                write_db_key=False,
                timestamp=timestamp,
            ))

        for row_number in sorted(approved_set):
            if row_number in exception_map:
                continue
            final_row = report.get_row(row_number)
            if final_row:
                self.apply_risk_metadata(final_row)
            if not final_row or not final_row.is_writable:
                skipped += 1
                continue
            if self._is_group_header(ws, row_number) or self._has_formula(ws, row_number, COL_VALUE):
                skipped += 1
                continue

            previous = ws.cell(row_number, COL_VALUE).value
            ws.cell(row_number, COL_VALUE).value = final_row.qc_value
            ws.cell(row_number, COL_VALUE).fill = (
                _FILL_MISMATCH if final_row.status == STATUS_MISMATCH else _FILL_APPROVED
            )
            if write_db_key:
                ws.cell(row_number, db_key_col).value = final_row.db_key

            applied += 1
            if final_row.source in ("manual", "manual_candidate"):
                learned_rows.append(final_row)

            log_entries.append(self._log_entry(
                row_number=row_number,
                final_row=final_row,
                previous=previous,
                status_after=STATUS_OK,
                reviewer_decision="Approved Correction",
                exception_reason="",
                write_db_key=write_db_key,
                timestamp=timestamp,
            ))

        self._write_log_sheet(wb, log_entries, report, applied, skipped)
        wb.save(str(output))

        learned_count = self.save_learned_mappings(report.model, learned_rows)
        post_report = self.analyze(str(output), report.qc_report, report.model)
        return ApplyResult(str(output), applied, skipped, learned_count, post_report, len(exception_map))

    def default_output_path(self, excel_path: str) -> str:
        path = Path(excel_path)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return str(path.with_name(f"{path.stem}_QC_Checked_AutoFilled_{timestamp}{path.suffix}"))

    def save_learned_mappings(self, model: str, rows: Iterable[FinalQcRow]) -> int:
        rows = [row for row in rows if row.db_key and row.module and row.item]
        if not rows:
            return 0

        existing = self._load_local_mappings(model)
        by_key = {
            (mapping.get("model", ""), normalize(mapping.get("module", "")), normalize(mapping.get("item_norm", ""))): mapping
            for mapping in existing
        }

        now = datetime.now().isoformat(timespec="seconds")
        for row in rows:
            key = (model, normalize(row.module), normalize(row.item))
            by_key[key] = {
                "model": model,
                "module": normalize(row.module),
                "item_norm": normalize(row.item),
                "db_key": row.db_key,
                "confidence": 0.95,
                "verified_at": now,
                "source": "reviewer_approved",
            }
            self._try_server_upsert(model, row)

        self._save_local_mappings(model, list(by_key.values()))
        return len(rows)

    def classify_value(self, current_value, qc_value) -> str:
        if self._is_empty_qc(qc_value):
            return STATUS_NON_COMPARABLE
        if self._is_empty_cell(current_value):
            return STATUS_MISSING
        return STATUS_OK if self._values_equal(current_value, qc_value) else STATUS_MISMATCH

    def apply_risk_metadata(self, row: FinalQcRow) -> FinalQcRow:
        """Attach correction risk and reviewer guidance to an analyzed row."""
        row.delta = self._format_delta(row.current_value, row.qc_value)

        if row.status in (STATUS_PROTECTED, STATUS_SKIPPED_GROUP, STATUS_UNMAPPED, STATUS_NON_COMPARABLE):
            row.risk_level = RISK_BLOCKED
            row.recommended_action = ACTION_DO_NOT_WRITE
            row.requires_review = False
            row.auto_approval_reason = row.protected_reason or "Not writable or not comparable"
            return row

        if row.status == STATUS_OK:
            row.risk_level = RISK_OK
            row.recommended_action = ACTION_NONE
            row.requires_review = False
            row.auto_approval_reason = "Checklist value already matches QC"
            return row

        if row.status not in (STATUS_MISSING, STATUS_MISMATCH) or not row.db_key:
            row.risk_level = RISK_BLOCKED
            row.recommended_action = ACTION_DO_NOT_WRITE
            row.requires_review = False
            row.auto_approval_reason = "No writable QC correction"
            return row

        if self._is_high_risk(row):
            row.risk_level = RISK_HIGH
            row.recommended_action = ACTION_REVIEW
            row.requires_review = True
            row.auto_approval_reason = "Large delta, type mismatch, or ambiguous candidates"
            return row

        if row.source in ("explicit", "learned", "exact"):
            row.risk_level = RISK_SAFE
            row.recommended_action = ACTION_FILL if row.status == STATUS_MISSING else ACTION_REPLACE
            row.requires_review = False
            row.auto_approval_reason = f"Trusted {row.source} mapping"
            return row

        row.risk_level = RISK_REVIEW
        row.recommended_action = ACTION_REVIEW
        row.requires_review = True
        row.auto_approval_reason = "Reviewer confirmation required"
        return row

    def _build_final_row(self, ws, raw: Dict, mapped, qc_lookup: Dict) -> FinalQcRow:
        row_num = raw["row"]
        current_value = ws.cell(row_num, COL_VALUE).value

        if self._is_group_header(ws, row_num):
            return self.apply_risk_metadata(FinalQcRow(
                row=row_num,
                module=raw["module"],
                item=raw["item"],
                current_value=current_value,
                qc_value=None,
                db_key=None,
                source="skipped",
                confidence=0.0,
                status=STATUS_SKIPPED_GROUP,
                protected_reason="Group header row",
            ))

        if self._has_formula(ws, row_num, COL_VALUE):
            qc_value = qc_lookup.get(mapped.db_key) if mapped.db_key else None
            return self.apply_risk_metadata(FinalQcRow(
                row=row_num,
                module=raw["module"],
                item=raw["item"],
                current_value=current_value,
                qc_value=qc_value,
                db_key=mapped.db_key,
                source=mapped.source,
                confidence=mapped.confidence,
                status=STATUS_PROTECTED,
                candidates=mapped.candidates,
                protected_reason="Measurement cell contains a formula",
            ))

        if not mapped.db_key:
            return self.apply_risk_metadata(FinalQcRow(
                row=row_num,
                module=raw["module"],
                item=raw["item"],
                current_value=current_value,
                qc_value=None,
                db_key=None,
                source=mapped.source,
                confidence=mapped.confidence,
                status=STATUS_UNMAPPED,
                candidates=mapped.candidates,
            ))

        qc_value = qc_lookup.get(mapped.db_key)
        status = self.classify_value(current_value, qc_value)
        return self.apply_risk_metadata(FinalQcRow(
            row=row_num,
            module=raw["module"],
            item=raw["item"],
            current_value=current_value,
            qc_value=qc_value,
            db_key=mapped.db_key,
            source=mapped.source,
            confidence=mapped.confidence,
            status=status,
            candidates=mapped.candidates,
        ))

    def _collect_rows(self, ws, db_key_col: int) -> List[Dict]:
        rows = []
        for row_num in range(2, ws.max_row + 1):
            module = ws.cell(row_num, COL_MODULE).value
            item = ws.cell(row_num, COL_ITEM).value
            if module is None and item is None:
                continue
            rows.append({
                "row": row_num,
                "module": str(module or "").strip(),
                "item": str(item or "").strip(),
                "unit": str(ws.cell(row_num, COL_UNIT).value or "").strip(),
                "explicit_db_key": str(ws.cell(row_num, db_key_col).value or "").strip(),
            })
        return rows

    def _find_main_sheet(self, wb):
        for name in wb.sheetnames:
            if fnmatch.fnmatch(name, "통합_*"):
                return wb[name]
        for name in wb.sheetnames:
            if name != "표지":
                return wb[name]
        return None

    def _find_db_key_column(self, ws) -> int:
        detected = self._validator._find_db_key_column(ws)
        return detected or COL_DB_KEY

    def _is_group_header(self, ws, row_num: int) -> bool:
        for col in (COL_MIN, COL_CRITERIA, COL_VALUE):
            if str(ws.cell(row_num, col).value or "").strip() == "-":
                return True
        return False

    def _has_formula(self, ws, row_num: int, col: int) -> bool:
        value = ws.cell(row_num, col).value
        return isinstance(value, str) and value.startswith("=")

    def _values_equal(self, left, right) -> bool:
        left_num = self._to_float(left)
        right_num = self._to_float(right)
        if left_num is not None and right_num is not None:
            return abs(left_num - right_num) <= self.tolerance
        return str(left).strip() == str(right).strip()

    def _is_empty_cell(self, value) -> bool:
        return value is None or str(value).strip() in ("", "-")

    def _is_empty_qc(self, value) -> bool:
        return value is None or str(value).strip() in ("", "N/A")

    def _to_float(self, value) -> Optional[float]:
        if value is None:
            return None
        text = str(value).strip().replace(",", "")
        if not text or text == "-":
            return None
        try:
            return float(text)
        except (TypeError, ValueError):
            return None

    def _is_high_risk(self, row: FinalQcRow) -> bool:
        if row.status != STATUS_MISMATCH:
            return False
        current_num = self._to_float(row.current_value)
        qc_num = self._to_float(row.qc_value)
        if (current_num is None) != (qc_num is None):
            return True
        if current_num is not None and qc_num is not None:
            delta = abs(current_num - qc_num)
            denominator = max(abs(qc_num), 1.0)
            if delta / denominator >= 0.50:
                return True
        return row.source in ("fuzzy", "unit_hint", "manual_candidate", "manual") and len(row.candidates) >= 2

    def _format_delta(self, current_value, qc_value) -> str:
        current_num = self._to_float(current_value)
        qc_num = self._to_float(qc_value)
        if current_num is not None and qc_num is not None:
            return f"{qc_num - current_num:+g}"
        if current_value is None or qc_value is None:
            return ""
        if current_num is None and qc_num is None:
            return ""
        return "type mismatch"

    def _extract_model_from_sheet(self, sheet_name: str) -> str:
        if "(" in sheet_name and ")" in sheet_name:
            start = sheet_name.find("(") + 1
            end = sheet_name.find(")", start)
            if end > start:
                return sheet_name[start:end].strip()
        return ""

    def _load_learned_mappings(self, model: str) -> List[Dict]:
        self._sync_remote_mappings(model)
        mappings = self._load_local_mappings(model)
        if self.sync_manager and hasattr(self.sync_manager, "_load_local_mappings"):
            try:
                mappings.extend(self.sync_manager._load_local_mappings(model))
            except Exception as exc:
                logger.warning(f"Unable to load sync-manager checklist mappings: {exc}")
        return mappings

    def _sync_remote_mappings(self, model: str):
        if self.sync_manager and hasattr(self.sync_manager, "sync_checklist_mappings"):
            connected_here = False
            try:
                if not getattr(self.sync_manager, "is_connected", False) and hasattr(self.sync_manager, "connect"):
                    connected_here = bool(self.sync_manager.connect())
                self.sync_manager.sync_checklist_mappings(model)
            except Exception as exc:
                logger.warning(f"Checklist mapping sync skipped: {exc}")
            finally:
                if connected_here and hasattr(self.sync_manager, "disconnect"):
                    self.sync_manager.disconnect()

    def _mapping_cache_path(self, model: str) -> Path:
        safe = "".join(ch if ch.isalnum() or ch in "-_ " else "_" for ch in (model or "default"))
        return self.mapping_dir / f"checklist_mappings_{safe}.json"

    def _load_local_mappings(self, model: str) -> List[Dict]:
        path = self._mapping_cache_path(model)
        if not path.exists():
            return []
        try:
            with open(path, "r", encoding="utf-8") as handle:
                return json.load(handle)
        except Exception as exc:
            logger.warning(f"Failed to load checklist mapping cache {path}: {exc}")
            return []

    def _save_local_mappings(self, model: str, mappings: List[Dict]):
        path = self._mapping_cache_path(model)
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w", encoding="utf-8") as handle:
            json.dump(mappings, handle, ensure_ascii=False, indent=2)

    def _try_server_upsert(self, model: str, row: FinalQcRow):
        if not self.sync_manager:
            return
        connected_here = False
        if not getattr(self.sync_manager, "is_connected", False) and hasattr(self.sync_manager, "connect"):
            connected_here = bool(self.sync_manager.connect())
        conn = getattr(self.sync_manager, "conn", None)
        if not conn:
            return
        try:
            from src.core.server_db_manager import ServerDBManager

            ServerDBManager(conn).upsert_checklist_mapping(
                model=model,
                module=normalize(row.module),
                item_norm=normalize(row.item),
                db_key=row.db_key,
                confidence=0.95,
                source="reviewer_approved",
            )
        except Exception as exc:
            logger.warning(f"Server checklist mapping upsert skipped: {exc}")
        finally:
            if connected_here and hasattr(self.sync_manager, "disconnect"):
                self.sync_manager.disconnect()

    def _check_output_directory(self, path: Path, result: PreflightResult):
        test_path = path.parent / ".__final_qc_preflight_write_test.tmp"
        try:
            with open(test_path, "w", encoding="utf-8") as handle:
                handle.write("ok")
            test_path.unlink(missing_ok=True)
        except Exception as exc:
            result.issues.append(
                PreflightIssue(
                    PREFLIGHT_ERROR,
                    "OUTPUT_DIRECTORY_WRITE",
                    f"Cannot write an output copy in this folder: {exc}",
                )
            )
            try:
                test_path.unlink(missing_ok=True)
            except Exception:
                pass

    def _check_required_columns(self, ws, result: PreflightResult):
        expectations = {
            COL_MODULE: ("MODULE_COLUMN", ("module",)),
            COL_ITEM: ("ITEM_COLUMN", ("check", "item")),
            COL_VALUE: ("MEASUREMENT_COLUMN", ("measurement", "value")),
            COL_UNIT: ("UNIT_COLUMN", ("unit",)),
        }
        for col, (code, keywords) in expectations.items():
            header = str(ws.cell(1, col).value or "").lower()
            if not header or not any(keyword in header for keyword in keywords):
                result.issues.append(
                    PreflightIssue(
                        PREFLIGHT_WARNING,
                        code,
                        f"Column {col} header does not look like the expected checklist field.",
                        sheet_name=ws.title,
                        column=col,
                    )
                )

    def _check_sheet_features(self, ws, result: PreflightResult, db_key_col: int):
        formula_rows = [
            row_num for row_num in range(2, ws.max_row + 1)
            if self._has_formula(ws, row_num, COL_VALUE)
        ]
        if formula_rows:
            result.issues.append(
                PreflightIssue(
                    PREFLIGHT_INFO,
                    "FORMULA_CELLS",
                    f"{len(formula_rows)} measurement cells contain formulas and will be blocked.",
                    sheet_name=ws.title,
                    column=COL_VALUE,
                )
            )

        hidden_rows = [idx for idx, dim in ws.row_dimensions.items() if dim.hidden and idx >= 2]
        if hidden_rows:
            result.issues.append(
                PreflightIssue(
                    PREFLIGHT_WARNING,
                    "HIDDEN_ROWS",
                    f"{len(hidden_rows)} hidden rows are present in the main sheet.",
                    sheet_name=ws.title,
                )
            )

        watched_cols = {COL_MODULE, COL_ITEM, COL_VALUE, COL_UNIT, db_key_col}
        merged_hits = []
        for merged_range in ws.merged_cells.ranges:
            if any(col in watched_cols for col in range(merged_range.min_col, merged_range.max_col + 1)):
                merged_hits.append(str(merged_range))
        if merged_hits:
            result.issues.append(
                PreflightIssue(
                    PREFLIGHT_WARNING,
                    "MERGED_CELLS",
                    f"Merged cells touch checklist/QC columns: {', '.join(merged_hits[:5])}",
                    sheet_name=ws.title,
                )
            )

        if ws.protection.sheet:
            result.issues.append(
                PreflightIssue(
                    PREFLIGHT_WARNING,
                    "SHEET_PROTECTION",
                    "The main checklist sheet is protected; writeback may fail.",
                    sheet_name=ws.title,
                )
            )

    def _log_entry(
        self,
        row_number: int,
        final_row: FinalQcRow,
        previous,
        status_after: str,
        reviewer_decision: str,
        exception_reason: str,
        write_db_key: bool,
        timestamp: str,
    ) -> List:
        return [
            row_number,
            final_row.status,
            status_after,
            final_row.risk_level,
            final_row.recommended_action,
            reviewer_decision,
            exception_reason,
            final_row.module,
            final_row.item,
            final_row.db_key,
            final_row.source,
            final_row.confidence,
            previous,
            final_row.qc_value,
            write_db_key,
            timestamp,
        ]

    def _write_log_sheet(self, wb, rows: List[List], report: FinalQcReport, applied: int, skipped: int):
        sheet_name = "_AutoFill_Log"
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws_log = wb.create_sheet(sheet_name)
        ws_log.append([
            "Row",
            "StatusBefore",
            "StatusAfter",
            "Risk",
            "RecommendedAction",
            "ReviewerDecision",
            "ExceptionReason",
            "Module",
            "Item",
            "DB_Key",
            "Source",
            "Confidence",
            "PreviousG",
            "QCValue",
            "WriteDBKey",
            "Timestamp",
        ])
        for entry in rows:
            ws_log.append(entry)
        ws_log.append([])
        ws_log.append(["Summary", f"Applied: {applied}", f"Skipped: {skipped}", f"Model: {report.model}"])
