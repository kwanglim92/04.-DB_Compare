"""
Checklist Validator Module
Compares Checklist Excel values against QC results.
DB_Key column in Excel drives the mapping — no external JSON needed.
"""

import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import openpyxl
import fnmatch

from src.constants import CHECKLIST_COLS
from src.utils.checklist_helpers import build_qc_lookup, find_db_key_column

logger = logging.getLogger(__name__)


class ChecklistValidationResult:
    """Result of a single checklist-DB comparison"""
    
    def __init__(self, checklist_text: str, checklist_row: int,
                 checklist_value, db_keys: List[str], db_values: Dict[str, any],
                 status: str, detail: str = ""):
        self.checklist_text = checklist_text
        self.checklist_row = checklist_row
        self.checklist_value = checklist_value
        self.db_keys = db_keys
        self.db_values = db_values  # {db_key: actual_value}
        self.status = status  # 'match', 'mismatch', 'missing_in_db'
        self.detail = detail


class ChecklistValidator:
    """Validates Checklist Excel against QC results using DB_Key column in Excel"""
    
    # Column indices — sourced from src.constants.CHECKLIST_COLS (single source of truth)
    ITEM_COL = CHECKLIST_COLS['ITEM']
    VALUE_COL = CHECKLIST_COLS['VALUE']
    MODULE_COL = CHECKLIST_COLS['MODULE']
    DB_KEY_COL = CHECKLIST_COLS['DB_KEY']
    
    def __init__(self):
        pass
    
    def validate(self, excel_path: str, qc_report: Dict) -> List[ChecklistValidationResult]:
        """
        Run full validation: parse Excel DB_Key column, compare with QC results.
        
        Args:
            excel_path: Path to the Checklist Excel file (must have DB_Key column)
            qc_report: QC inspection report from QCComparator.generate_report()
            
        Returns:
            List of ChecklistValidationResult
        """
        results = []
        
        # Parse Excel — find DB_Key column and extract mapped rows
        mapped_rows = self._parse_excel(excel_path)
        if not mapped_rows:
            logger.warning("No DB_Key mappings found in Excel")
            return results
        
        # Build lookup from QC results
        qc_lookup = self._build_qc_lookup(qc_report)
        
        # Compare each mapped row
        for row in mapped_rows:
            db_keys = row['db_keys']  # List of keys (semicolon-split)
            checklist_value = row['value']
            checklist_text = row['item']
            row_num = row['row']
            
            # Get QC values for all mapped keys
            db_values = {}
            for key in db_keys:
                val = qc_lookup.get(key.strip())
                db_values[key.strip()] = val
            
            # All keys missing from QC results?
            if all(v is None for v in db_values.values()):
                results.append(ChecklistValidationResult(
                    checklist_text=checklist_text,
                    checklist_row=row_num,
                    checklist_value=checklist_value,
                    db_keys=db_keys,
                    db_values=db_values,
                    status='missing_in_db',
                    detail="DB key not found in QC results"
                ))
                continue
            
            # Compare values
            status, detail = self._compare_values(checklist_value, db_values)
            
            results.append(ChecklistValidationResult(
                checklist_text=checklist_text,
                checklist_row=row_num,
                checklist_value=checklist_value,
                db_keys=db_keys,
                db_values=db_values,
                status=status,
                detail=detail
            ))
        
        # Log summary
        match_count = sum(1 for r in results if r.status == 'match')
        mismatch_count = sum(1 for r in results if r.status == 'mismatch')
        missing_count = sum(1 for r in results if r.status == 'missing_in_db')
        logger.info(f"Validation complete: {match_count} match, {mismatch_count} mismatch, {missing_count} missing")
        
        return results
    
    def _parse_excel(self, excel_path: str) -> List[Dict]:
        """Parse Excel: find DB_Key column and extract rows that have DB_Key values"""
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            
            # Find the checklist sheet (pattern: 통합_*)
            ws = None
            for name in wb.sheetnames:
                if fnmatch.fnmatch(name, '통합_*'):
                    ws = wb[name]
                    break
            
            if ws is None:
                ws = wb.active
                logger.warning(f"Sheet '통합_*' not found, using active sheet: {ws.title}")
            
            # Auto-detect DB_Key column from header row
            db_key_col = self._find_db_key_column(ws)
            if db_key_col is None:
                logger.warning("DB_Key column not found in Excel header")
                return []
            
            logger.info(f"Found DB_Key column at index {db_key_col} (sheet: {ws.title})")
            
            # Extract rows where DB_Key has a value
            mapped_rows = []
            for r in range(2, ws.max_row + 1):
                db_key_value = ws.cell(r, db_key_col).value
                if not db_key_value:
                    continue
                
                # Split multiple keys by semicolon
                db_key_str = str(db_key_value).strip()
                db_keys = [k.strip() for k in db_key_str.split(';') if k.strip()]
                
                if not db_keys:
                    continue
                
                mapped_rows.append({
                    'row': r,
                    'module': ws.cell(r, self.MODULE_COL).value or '',
                    'item': str(ws.cell(r, self.ITEM_COL).value or '').strip(),
                    'value': ws.cell(r, self.VALUE_COL).value,
                    'db_keys': db_keys
                })
            
            logger.info(f"Found {len(mapped_rows)} rows with DB_Key mappings")
            return mapped_rows
            
        except Exception as e:
            logger.error(f"Excel parse error: {e}", exc_info=True)
            return []
    
    def _find_db_key_column(self, ws) -> Optional[int]:
        """Delegate to src.utils.checklist_helpers.find_db_key_column (kept for backward compat)."""
        return find_db_key_column(ws)

    def _build_qc_lookup(self, qc_report: Dict) -> Dict[str, any]:
        """Delegate to src.utils.checklist_helpers.build_qc_lookup (kept for backward compat)."""
        return build_qc_lookup(qc_report)


    def _compare_values(self, checklist_value, db_values: Dict) -> Tuple[str, str]:
        """Compare checklist value with DB values (numeric comparison)"""
        if checklist_value is None or str(checklist_value).strip() in ('', '-', 'N/A', 'Yes', 'No'):
            return ('match', 'Non-numeric value, skipped')
        
        # Parse checklist value as number
        try:
            cl_num = float(str(checklist_value))
        except (ValueError, TypeError):
            return ('match', f'Non-numeric checklist value: {checklist_value}')
        
        # Compare against each DB key value
        mismatches = []
        matches = []
        
        for db_key, db_val in db_values.items():
            if db_val is None:
                continue
            
            try:
                db_num = float(str(db_val))
            except (ValueError, TypeError):
                mismatches.append(f"{db_key}: DB value '{db_val}' is not numeric")
                continue
            
            if abs(cl_num - db_num) < 0.001:
                matches.append(f"{db_key}: {cl_num} == {db_num}")
            else:
                mismatches.append(f"{db_key}: checklist={cl_num}, DB={db_num}")
        
        if mismatches:
            return ('mismatch', '; '.join(mismatches))
        elif matches:
            return ('match', '; '.join(matches))
        else:
            return ('match', 'No comparable DB values')
    
    @staticmethod
    def export_db_keys(qc_report: Dict, filepath: str) -> bool:
        """Export all DB keys from QC results to a text file for user reference"""
        try:
            lines = [
                "# DB Keys — Copy the key and paste into your Checklist's DB_Key column",
                "# Format: Module.PartType.PartName.ItemName = actual_value",
                "#" + "=" * 80,
                ""
            ]
            
            # Group by module
            modules = {}
            for result in qc_report.get('results', []):
                mod = result['module']
                if mod not in modules:
                    modules[mod] = []
                key = f"{mod}.{result['part_type']}.{result['part_name']}.{result['item_name']}"
                val = result.get('actual_value', '')
                modules[mod].append((key, val))
            
            for mod_name in sorted(modules.keys()):
                lines.append(f"### {mod_name} ###")
                for key, val in sorted(modules[mod_name]):
                    lines.append(f"  {key} = {val}")
                lines.append("")
            
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write('\n'.join(lines))
            
            logger.info(f"Exported {sum(len(v) for v in modules.values())} DB keys to {filepath}")
            return True
            
        except Exception as e:
            logger.error(f"Export DB keys failed: {e}")
            return False
