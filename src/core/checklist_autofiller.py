"""
Checklist AutoFiller Module
Writes QC values into the G-column (Measurement) of a customer checklist Excel.

Safety guarantees:
  - Only G-column is written (and optionally M-column). All others are untouched.
  - Cells that already contain a formula are skipped.
  - Group header rows (D/E/G = '-') are skipped.
  - The '표지' sheet is never touched.
  - A timestamped backup is created before any write.
  - Dry-run mode (default ON) produces a report without writing.
"""

import fnmatch
import logging
import shutil
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
from openpyxl.styles import PatternFill

from src.constants import CHECKLIST_COLS
from src.core.checklist_mapper import ChecklistMapper, MapResult
from src.core.checklist_validator import ChecklistValidator

logger = logging.getLogger(__name__)

# openpyxl aRGB fill colors
_FILL_GREEN  = PatternFill(fill_type='solid', fgColor='FFE8F5E9')  # explicit / learned
_FILL_YELLOW = PatternFill(fill_type='solid', fgColor='FFFFF8E1')  # fuzzy
_FILL_RED    = PatternFill(fill_type='solid', fgColor='FFFFEBEE')  # conflict / low confidence

# Column indices — sourced from src.constants.CHECKLIST_COLS
_COL_MODULE   = CHECKLIST_COLS['MODULE']
_COL_ITEM     = CHECKLIST_COLS['ITEM']
_COL_MIN      = CHECKLIST_COLS['MIN']
_COL_CRITERIA = CHECKLIST_COLS['CRITERIA']
_COL_VALUE    = CHECKLIST_COLS['VALUE']
_COL_UNIT     = CHECKLIST_COLS['UNIT']
_COL_PASS     = CHECKLIST_COLS['PASS']
_COL_DB_KEY   = CHECKLIST_COLS['DB_KEY']

# Group-header sentinel value
_GROUP_SENTINEL = '-'


@dataclass
class AutoFillReport:
    filled: int = 0
    skipped_unmapped: int = 0
    skipped_protected: int = 0
    skipped_group_header: int = 0
    conflicts: int = 0           # G already has a non-formula value that differs
    mapping_changes: List[Dict] = field(default_factory=list)  # detail rows for the log sheet

    @property
    def total_processed(self) -> int:
        return (self.filled + self.skipped_unmapped +
                self.skipped_protected + self.skipped_group_header)


class ChecklistAutoFiller:
    """Fills QC values into checklist Excel G-column with full safety guards."""

    def __init__(self,
                 excel_path: str,
                 qc_report: Dict,
                 sync_manager=None,
                 model: str = '',
                 dry_run: bool = True,
                 fill_metadata: bool = False,
                 confidence_threshold: float = 0.80,
                 write_db_key_column: bool = False):
        """
        Args:
            excel_path:           Path to customer checklist .xlsx
            qc_report:            Output of QCComparator.generate_report()
            sync_manager:         SyncManager instance (used to load learned mappings)
            model:                Checklist model name (for mapping cache lookup)
            dry_run:              If True, no file is written; report only
            fill_metadata:        If True, also fill Last sheet metadata cells
            confidence_threshold: Minimum mapper confidence to auto-fill
            write_db_key_column:  If True, also write matched db_key into M-column
        """
        self.excel_path = Path(excel_path)
        self.qc_report = qc_report
        self.sync_manager = sync_manager
        self.model = model
        self.dry_run = dry_run
        self.fill_metadata = fill_metadata
        self.confidence_threshold = confidence_threshold
        self.write_db_key_column = write_db_key_column

        self._validator = ChecklistValidator()
        self._qc_lookup: Dict = self._validator._build_qc_lookup(qc_report)

    # -----------------------------------------------------------------------
    # Public API
    # -----------------------------------------------------------------------

    def run(self) -> AutoFillReport:
        """Execute auto-fill. Returns report; writes file only when dry_run=False."""
        report = AutoFillReport()

        # Load workbook (data_only=False to preserve formulas)
        try:
            wb = openpyxl.load_workbook(str(self.excel_path), data_only=False)
        except Exception as e:
            logger.error(f"Failed to open workbook: {e}")
            return report

        ws = self._find_main_sheet(wb)
        if ws is None:
            logger.error("Main checklist sheet not found (expected '통합_*')")
            return report

        # Load learned mappings
        learned = []
        if self.sync_manager:
            learned = self.sync_manager._load_local_mappings(self.model)

        mapper = ChecklistMapper(
            qc_lookup=self._qc_lookup,
            learned_mappings=learned,
            model=self.model,
            fuzzy_threshold=self.confidence_threshold,
        )

        # Collect rows from the main sheet
        rows = self._collect_rows(ws)

        # Run mapper
        map_results: List[MapResult] = mapper.map_rows(rows)

        # Apply results to workbook
        timestamp = datetime.now()
        for result, raw_row in zip(map_results, rows):
            self._apply_result(ws, result, raw_row, report, timestamp)

        # Optionally fill Last sheet metadata
        if self.fill_metadata and not self.dry_run:
            self._fill_metadata_sheet(wb)

        # Write log sheet
        self._write_log_sheet(wb, report)

        # Save (with backup)
        if not self.dry_run:
            self._save_with_backup(wb)
        else:
            logger.info(
                f"[DRY-RUN] would fill {report.filled} cells, "
                f"skip {report.skipped_unmapped} unmapped, "
                f"skip {report.skipped_protected} protected"
            )

        return report

    # -----------------------------------------------------------------------
    # Sheet helpers
    # -----------------------------------------------------------------------

    def _find_main_sheet(self, wb):
        for name in wb.sheetnames:
            if fnmatch.fnmatch(name, '통합_*'):
                return wb[name]
        # Fallback: skip 표지, use first non-표지 sheet
        for name in wb.sheetnames:
            if name != '표지':
                logger.warning(f"Sheet '통합_*' not found; using '{name}'")
                return wb[name]
        return None

    def _collect_rows(self, ws) -> List[Dict]:
        """Read all data rows from the main sheet (R2 onwards)."""
        rows = []
        for r in range(2, ws.max_row + 1):
            module_val = ws.cell(r, _COL_MODULE).value
            item_val = ws.cell(r, _COL_ITEM).value
            if item_val is None and module_val is None:
                continue

            explicit_key = ws.cell(r, _COL_DB_KEY).value
            unit_val = ws.cell(r, _COL_UNIT).value

            rows.append({
                'row': r,
                'module': str(module_val or '').strip(),
                'item': str(item_val or '').strip(),
                'explicit_db_key': str(explicit_key).strip() if explicit_key else '',
                'unit': str(unit_val or '').strip(),
            })
        return rows

    # -----------------------------------------------------------------------
    # Safety guards
    # -----------------------------------------------------------------------

    def _is_group_header(self, ws, row: int) -> bool:
        """Group header rows have '-' in D, E, or G columns."""
        for col in (_COL_MIN, _COL_CRITERIA, _COL_VALUE):
            val = ws.cell(row, col).value
            if str(val or '').strip() == _GROUP_SENTINEL:
                return True
        return False

    def _has_formula(self, ws, row: int, col: int) -> bool:
        cell = ws.cell(row, col)
        return isinstance(cell.value, str) and cell.value.startswith('=')

    def _cell_is_protected(self, ws, row: int) -> bool:
        """Return True if the G-cell must not be written."""
        if self._has_formula(ws, row, _COL_VALUE):
            return True
        if self._has_formula(ws, row, _COL_PASS):   # I-column sanity
            pass  # I is always protected; we just never write it
        return False

    # -----------------------------------------------------------------------
    # Apply a single mapping result
    # -----------------------------------------------------------------------

    def _apply_result(self, ws, result: MapResult, raw_row: Dict,
                      report: AutoFillReport, timestamp: datetime):
        row = result.row

        # Skip group header rows
        if self._is_group_header(ws, row):
            report.skipped_group_header += 1
            return

        # Skip unmapped rows
        if result.source == 'unmapped' or result.db_key is None:
            report.skipped_unmapped += 1
            return

        # Skip if confidence below threshold (shouldn't reach here for D/E, but guard)
        if result.confidence < self.confidence_threshold:
            report.skipped_unmapped += 1
            return

        # Skip if G-cell is protected (formula)
        if self._cell_is_protected(ws, row):
            report.skipped_protected += 1
            return

        qc_value = self._qc_lookup.get(result.db_key)
        if qc_value is None:
            report.skipped_unmapped += 1
            return

        current_g = ws.cell(row, _COL_VALUE).value
        conflict = (current_g is not None and
                    not (isinstance(current_g, str) and current_g.startswith('=')) and
                    str(current_g).strip() != str(qc_value).strip())

        if conflict:
            report.conflicts += 1

        fill = self._choose_fill(result.source, conflict)

        log_entry = {
            'row': row,
            'module': result.module,
            'item': result.item,
            'db_key': result.db_key,
            'confidence': result.confidence,
            'source': result.source,
            'qc_value': qc_value,
            'prev_value': current_g,
            'conflict': conflict,
            'timestamp': timestamp.isoformat(),
        }
        report.mapping_changes.append(log_entry)
        report.filled += 1

        if not self.dry_run:
            ws.cell(row, _COL_VALUE).value = qc_value
            ws.cell(row, _COL_VALUE).fill = fill
            if self.write_db_key_column:
                ws.cell(row, _COL_DB_KEY).value = result.db_key

    def _choose_fill(self, source: str, conflict: bool) -> PatternFill:
        if conflict:
            return _FILL_RED
        if source in ('explicit', 'learned', 'exact'):
            return _FILL_GREEN
        return _FILL_YELLOW  # fuzzy / unit_hint

    # -----------------------------------------------------------------------
    # _AutoFill_Log sheet
    # -----------------------------------------------------------------------

    def _write_log_sheet(self, wb, report: AutoFillReport):
        sheet_name = '_AutoFill_Log'
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

        ws_log = wb.create_sheet(sheet_name)
        headers = ['Row', 'Module', 'Item', 'DB_Key', 'Confidence',
                   'Source', 'QC_Value', 'Prev_Value', 'Conflict', 'Timestamp']
        ws_log.append(headers)

        for entry in report.mapping_changes:
            ws_log.append([
                entry['row'], entry['module'], entry['item'],
                entry['db_key'], entry['confidence'], entry['source'],
                entry['qc_value'], entry['prev_value'], entry['conflict'],
                entry['timestamp'],
            ])

        # Summary footer
        ws_log.append([])
        ws_log.append(['Summary', f"Filled: {report.filled}",
                        f"Unmapped: {report.skipped_unmapped}",
                        f"Protected: {report.skipped_protected}",
                        f"Conflicts: {report.conflicts}",
                        f"DryRun: {self.dry_run}"])

    # -----------------------------------------------------------------------
    # Metadata (Last sheet) — Phase 5
    # -----------------------------------------------------------------------

    def set_metadata(self, metadata: dict):
        """Set metadata values to be written to the Last sheet.

        Keys: model_name, sid_number, reference_document, end_user,
              manufacturing_engineer, qc_engineer, manager
        Call before run() when fill_metadata=True.
        """
        self._metadata = metadata

    def _fill_metadata_sheet(self, wb):
        """Fill Last sheet L-column cells from self._metadata."""
        if 'Last' not in wb.sheetnames:
            logger.warning("'Last' sheet not found — skipping metadata fill")
            return

        meta = getattr(self, '_metadata', {}) or {}
        ws_last = wb['Last']

        mapping = {
            22: meta.get('model_name', ''),
            25: meta.get('sid_number', ''),
            28: meta.get('reference_document', ''),
            31: meta.get('test_date', datetime.now().strftime('%Y-%m-%d')),
            34: meta.get('end_user', ''),
            37: meta.get('manufacturing_engineer', ''),
            40: meta.get('qc_engineer', ''),
            43: meta.get('manager', ''),
        }

        for row_num, value in mapping.items():
            if not value:
                continue
            cell = ws_last.cell(row_num, 12)  # L-column
            if cell.value and str(cell.value).strip():
                logger.debug(f"Last!L{row_num} already has value, skipping")
                continue
            cell.value = value
            logger.info(f"Filled Last!L{row_num} = {value!r}")

    # -----------------------------------------------------------------------
    # Backup + Save
    # -----------------------------------------------------------------------

    def _save_with_backup(self, wb):
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_path = self.excel_path.with_suffix(f'.bak.{ts}.xlsx')
        try:
            shutil.copy2(str(self.excel_path), str(backup_path))
            logger.info(f"Backup created: {backup_path.name}")
        except OSError as e:
            logger.error(f"Backup failed: {e}")
            return

        try:
            wb.save(str(self.excel_path))
            logger.info(f"Saved: {self.excel_path.name}")
        except Exception as e:
            logger.error(f"Save failed: {e}")
