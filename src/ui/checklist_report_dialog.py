"""
Checklist Validation Report Dialog
Shows cross-check results between Checklist Excel and DB values.
"""

import customtkinter as ctk
from tkinter import ttk, filedialog, messagebox
import logging
from datetime import datetime
from pathlib import Path

logger = logging.getLogger(__name__)


class ChecklistReportDialog(ctk.CTkToplevel):
    """Dialog showing checklist validation results with mismatch report"""
    
    def __init__(self, parent, results, excel_path: str, profile_name: str, qc_report: dict = None):
        super().__init__(parent)
        
        self.results = results
        self.excel_path = excel_path
        self.profile_name = profile_name
        self.qc_report = qc_report
        
        # Window setup
        self.title("📋 Checklist Validation Report")
        self.geometry("1000x650")
        self.transient(parent)
        self.grab_set()
        
        self.create_widgets()
    
    def create_widgets(self):
        # === Header ===
        header_frame = ctk.CTkFrame(self, fg_color="transparent")
        header_frame.pack(fill="x", padx=20, pady=(15, 5))
        
        ctk.CTkLabel(
            header_frame,
            text="Checklist Validation Report",
            font=("Segoe UI", 18, "bold")
        ).pack(anchor="w")
        
        # File info
        excel_name = Path(self.excel_path).name
        ctk.CTkLabel(
            header_frame,
            text=f"File: {excel_name}  |  Profile: {self.profile_name}",
            font=("Segoe UI", 11),
            text_color="#888888"
        ).pack(anchor="w", pady=(2, 0))
        
        # === Summary Stats ===
        stats_frame = ctk.CTkFrame(self)
        stats_frame.pack(fill="x", padx=20, pady=10)
        
        total = len(self.results)
        match_count = sum(1 for r in self.results if r.status == 'match')
        mismatch_count = sum(1 for r in self.results if r.status == 'mismatch')
        missing_cl = sum(1 for r in self.results if r.status == 'missing_in_checklist')
        missing_db = sum(1 for r in self.results if r.status == 'missing_in_db')
        missing_count = missing_cl + missing_db
        
        # Error rate
        comparable = match_count + mismatch_count
        error_rate = (mismatch_count / comparable * 100) if comparable > 0 else 0
        
        stats_inner = ctk.CTkFrame(stats_frame, fg_color="transparent")
        stats_inner.pack(pady=10, padx=15)
        
        # Stats labels
        stats_data = [
            ("Checked", f"{total}", "#4a9eff"),
            ("✅ Match", f"{match_count}", "#27ae60"),
            ("🔴 Mismatch", f"{mismatch_count}", "#e74c3c"),
            ("⚠ Missing", f"{missing_count}", "#f39c12"),
            ("오기입율", f"{error_rate:.1f}%", "#e74c3c" if error_rate > 0 else "#27ae60"),
        ]
        
        for i, (label, value, color) in enumerate(stats_data):
            col_frame = ctk.CTkFrame(stats_inner, fg_color="transparent")
            col_frame.grid(row=0, column=i, padx=20)
            
            ctk.CTkLabel(
                col_frame, text=value,
                font=("Segoe UI", 22, "bold"),
                text_color=color
            ).pack()
            
            ctk.CTkLabel(
                col_frame, text=label,
                font=("Segoe UI", 11),
                text_color="#aaaaaa"
            ).pack()
        
        # === Filter toggle ===
        filter_frame = ctk.CTkFrame(self, fg_color="transparent")
        filter_frame.pack(fill="x", padx=20, pady=(5, 0))
        
        self.show_all_var = ctk.BooleanVar(value=False)
        ctk.CTkCheckBox(
            filter_frame,
            text="Show all items (including matches)",
            variable=self.show_all_var,
            command=self._refresh_table,
            font=("Segoe UI", 11)
        ).pack(side="left")
        
        # === Bottom buttons (pack BEFORE table so they always show) ===
        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.pack(fill="x", padx=20, pady=(0, 15), side="bottom")
        
        ctk.CTkButton(
            btn_frame, text="Export to Excel",
            command=self._export_excel,
            font=("Segoe UI", 13),
            fg_color="#2196F3", hover_color="#1976D2",
            width=160, height=38
        ).pack(side="left", padx=(0, 10))
        
        ctk.CTkButton(
            btn_frame, text="Export DB Keys",
            command=self._export_db_keys,
            font=("Segoe UI", 13),
            fg_color="#7b1fa2", hover_color="#5c1680",
            width=150, height=38
        ).pack(side="left")
        
        ctk.CTkButton(
            btn_frame, text="Close",
            command=self.destroy,
            font=("Segoe UI", 13),
            fg_color="#555555", hover_color="#666666",
            width=100, height=38
        ).pack(side="right")
        
        # === Results Table ===
        table_frame = ctk.CTkFrame(self)
        table_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        columns = ('row', 'status', 'checklist_item', 'checklist_val', 'db_val', 'detail')
        self.tree = ttk.Treeview(
            table_frame, columns=columns, show='headings', height=15
        )
        
        self.tree.heading('row', text='Row')
        self.tree.heading('status', text='Status')
        self.tree.heading('checklist_item', text='Check Item')
        self.tree.heading('checklist_val', text='Checklist Value')
        self.tree.heading('db_val', text='DB Value')
        self.tree.heading('detail', text='Detail')
        
        self.tree.column('row', width=50, anchor='center')
        self.tree.column('status', width=80, anchor='center')
        self.tree.column('checklist_item', width=250)
        self.tree.column('checklist_val', width=100, anchor='center')
        self.tree.column('db_val', width=100, anchor='center')
        self.tree.column('detail', width=300)
        
        # Tags
        self.tree.tag_configure('match', foreground='#27ae60')
        self.tree.tag_configure('mismatch', foreground='#e74c3c', background='#3a2020')
        self.tree.tag_configure('missing', foreground='#f39c12')
        
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Style
        style = ttk.Style()
        style.configure("Treeview", rowheight=28, font=("Segoe UI", 11))
        style.configure("Treeview.Heading", font=("Segoe UI", 11, "bold"))
        
        # Initial populate
        self._refresh_table()
    
    def _refresh_table(self):
        """Populate/refresh the results table"""
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        show_all = self.show_all_var.get()
        
        for r in self.results:
            if not show_all and r.status == 'match':
                continue
            
            # Status icon
            if r.status == 'match':
                status_text = '✅ OK'
                tag = 'match'
            elif r.status == 'mismatch':
                status_text = '🔴 ERR'
                tag = 'mismatch'
            else:
                status_text = '⚠ N/A'
                tag = 'missing'
            
            # DB value: show first non-None value
            db_display = ''
            for key, val in r.db_values.items():
                if val is not None:
                    db_display = str(val)
                    break
            
            self.tree.insert('', 'end', values=(
                r.checklist_row if r.checklist_row > 0 else '-',
                status_text,
                r.checklist_text[:50],
                r.checklist_value if r.checklist_value is not None else '-',
                db_display or '-',
                r.detail[:60]
            ), tags=(tag,))
    
    def _export_excel(self):
        """Export validation results to Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            default_name = f"Checklist_Validation_{self.profile_name}_{timestamp}.xlsx"
            
            filepath = filedialog.asksaveasfilename(
                title="Export Validation Report",
                defaultextension=".xlsx",
                initialfile=default_name,
                filetypes=[("Excel files", "*.xlsx")]
            )
            
            if not filepath:
                return
            
            wb = Workbook()
            
            # === Summary Sheet ===
            ws_summary = wb.active
            ws_summary.title = "Summary"
            
            total = len(self.results)
            match_count = sum(1 for r in self.results if r.status == 'match')
            mismatch_count = sum(1 for r in self.results if r.status == 'mismatch')
            missing_count = sum(1 for r in self.results if r.status.startswith('missing'))
            comparable = match_count + mismatch_count
            error_rate = (mismatch_count / comparable * 100) if comparable > 0 else 0
            
            ws_summary.append(["Checklist Validation Report"])
            ws_summary.append([""])
            ws_summary.append(["Source File", Path(self.excel_path).name])
            ws_summary.append(["Profile", self.profile_name])
            ws_summary.append(["Date", datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
            ws_summary.append([""])
            ws_summary.append(["Total Checked", total])
            ws_summary.append(["Match", match_count])
            ws_summary.append(["Mismatch", mismatch_count])
            ws_summary.append(["Missing", missing_count])
            ws_summary.append(["Error Rate", f"{error_rate:.1f}%"])
            
            # Bold header
            ws_summary['A1'].font = Font(bold=True, size=14)
            
            # === Detail Sheet ===
            ws_detail = wb.create_sheet("Details")
            
            headers = ["Row", "Status", "Check Item", "Checklist Value", "DB Value", "DB Key", "Detail"]
            ws_detail.append(headers)
            
            # Header style
            header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for col in range(1, len(headers) + 1):
                cell = ws_detail.cell(1, col)
                cell.fill = header_fill
                cell.font = header_font
            
            # Data rows
            red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            
            for r in self.results:
                db_display = ''
                for key, val in r.db_values.items():
                    if val is not None:
                        db_display = str(val)
                        break
                
                status_text = {
                    'match': 'OK',
                    'mismatch': 'MISMATCH',
                    'missing_in_checklist': 'MISSING (CL)',
                    'missing_in_db': 'MISSING (DB)'
                }.get(r.status, r.status)
                
                row_data = [
                    r.checklist_row if r.checklist_row > 0 else '-',
                    status_text,
                    r.checklist_text,
                    r.checklist_value if r.checklist_value is not None else '-',
                    db_display or '-',
                    ', '.join(r.db_keys),
                    r.detail
                ]
                ws_detail.append(row_data)
                
                # Highlight mismatches
                if r.status == 'mismatch':
                    row_num = ws_detail.max_row
                    for col in range(1, len(headers) + 1):
                        ws_detail.cell(row_num, col).fill = red_fill
            
            # Auto-fit columns
            for col_cells in ws_detail.columns:
                max_length = max(len(str(cell.value or '')) for cell in col_cells)
                ws_detail.column_dimensions[col_cells[0].column_letter].width = min(max_length + 4, 50)
            
            wb.save(filepath)
            messagebox.showinfo("Success", f"Report exported:\n{filepath}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Export failed:\n{e}")
            logger.error(f"Export error: {e}", exc_info=True)
    
    def _export_db_keys(self):
        """Export all available DB keys from QC results"""
        if not self.qc_report:
            messagebox.showwarning("Warning", "No QC report data available.")
            return
        
        try:
            from src.core.checklist_validator import ChecklistValidator
            
            filepath = filedialog.asksaveasfilename(
                title="Export DB Keys",
                defaultextension=".txt",
                initialfile=f"DB_Keys_{self.profile_name}.txt",
                filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
            )
            
            if not filepath:
                return
            
            success = ChecklistValidator.export_db_keys(self.qc_report, filepath)
            if success:
                messagebox.showinfo(
                    "Success",
                    f"DB Keys exported to:\n{filepath}\n\n"
                    "Copy the keys and paste into your Checklist's DB_Key column (M열)."
                )
            else:
                messagebox.showerror("Error", "Failed to export DB keys.")
                
        except Exception as e:
            messagebox.showerror("Error", f"Export DB keys failed:\n{e}")
            logger.error(f"Export DB keys error: {e}", exc_info=True)
